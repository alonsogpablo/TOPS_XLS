import os
from os.path import join
import csv

from openpyxl import Workbook
from openpyxl import formatting, styles
from openpyxl.styles import colors
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

from os import listdir
from os.path import isfile, join



path='C:\\Users\\palonso0\\PycharmProjects\\TOPS_XLS\\estad'

# Create fill
redFill = PatternFill(start_color='FF99CC', end_color='FF99CC')


onlyfiles = [f for f in listdir(path) if (isfile(join(path, f)) and f.endswith('.csv') and (f.startswith('eri') or f.startswith('hua')or f.startswith('top')))]


sheet_names=[]
for file in onlyfiles:
    name=file[:-4]
    sheet_names.append(name)

wb=Workbook()
del wb['Sheet']

#Create INDEX SHEET

wb.create_sheet('INDICE')
ws=wb['INDICE']

i=1
for f in onlyfiles:

    link = '#'+f[:-4]+'!A1'
    ws.cell(row=i,column=1).value=f[:-4]
    ws.cell(row=i, column=1).font = Font(color=colors.BLUE)
    ws.cell(row=i, column=1).hyperlink = (link)
    i=i+2

ws.sheet_view.zoomScale = 75

for sheet in sheet_names:
    wb.create_sheet(sheet)

for f in onlyfiles:
    src=csv.reader(open(path+'\\'+f),delimiter=';')
    ws=wb[f[:-4]]
    for row_index,row in enumerate(src):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            s=cell
            try:
                s=float(s)
                ws[('%s%s' % (column_letter, (row_index + 1)))].value = s

            except ValueError:
                ws[('%s%s' % (column_letter, (row_index + 1)))].value = s

            if f[:-4]<>'top_Hua_VoLTE':
                link = '#INDICE!A1'
                ws.cell(row=1, column=12).value = 'INDICE'
                ws.cell(row=1, column=12).font = Font(color=colors.BLUE)
                ws.cell(row=1, column=12).hyperlink = (link)
            elif f[:-4]=='top_Hua_VoLTE':
                link = '#INDICE!A1'
                ws.cell(row=1, column=22).value = 'INDICE'
                ws.cell(row=1, column=22).font = Font(color=colors.BLUE)
                ws.cell(row=1, column=22).hyperlink = (link)


        ws.sheet_view.zoomScale = 75

ws=wb['eri3gdia_top_rssi']
ws.conditional_formatting.add('E1:E300',CellIsRule(operator='greaterThan', formula=['-90'], stopIfTrue=True, fill=redFill))

ws=wb['eri4gdia_top_interf']
ws.conditional_formatting.add('E1:E300',CellIsRule(operator='greaterThan', formula=['-100'], stopIfTrue=True, fill=redFill))

ws=wb['hua4gdia_top_interf']
ws.conditional_formatting.add('E1:E300',CellIsRule(operator='greaterThan', formula=['-100'], stopIfTrue=True, fill=redFill))

ws=wb['hua3gdia_top_rssi']
ws.conditional_formatting.add('E1:E300',CellIsRule(operator='greaterThan', formula=['-90'], stopIfTrue=True, fill=redFill))


wb.active=0

wb.save('TOPS.xlsx')